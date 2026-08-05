"""
Parses the manufacturer's order-form Excel file (e.g. North Kiteboarding
order forms) into a clean, normalized list of product rows.

The manufacturer's file has a fixed shape we've observed across seasons:
  - A cover area (rows ~1-11) with totals/discounts - not product data.
  - A header row (found dynamically) with columns such as:
    Ranking, Item Sub Group Description, Segment, Item, Barcode (Item),
    Item Description, Color, Color Description, Size, New / Carry over,
    Suggested Retail Price, USDEX Price, Order, Order amount, ...
  - Product rows below the header, one per SKU (item + color + size).

There is usually one sheet (e.g. "North ALL products") that already
consolidates every category, alongside category-specific sheets with the
same column layout. By default we read every sheet and concatenate, since
category sheets are a subset of the "ALL" sheet in some files and the
"ALL" sheet may be absent or named differently in others; callers can
restrict to a specific sheet name if they want to avoid double counting.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl

# Column headers we look for, and the normalized field name we map them to.
# Matching is case-insensitive and ignores leading/trailing whitespace, since
# the manufacturer's own files are inconsistent about this (e.g. "Item " vs
# "Item", "New / Carry over " vs "New/Carry Over").
_HEADER_ALIASES: dict[str, str] = {
    "ranking": "ranking",
    "item sub group description": "sub_group",
    "segment": "segment",
    "item": "item_code",
    "barcode (item)": "barcode",
    "barcode": "barcode",
    "item description": "description",
    "color": "color_code",
    "color description": "color_description",
    "size": "size",
    "new / carry over": "status",
    "new/carry over": "status",
    "status": "status",  # seen on Mystic order forms (North's is "New / Carry over")
    "suggested retail price": "retail_price",
    "suggested retail price (ex vat)": "retail_price",  # seen on Mystic USD order forms
    "usdex price": "wholesale_price",
    "standard sales price": "wholesale_price",  # Mystic's name for the same kind of field
    "order": "order_qty",
    "order amount": "order_amount",
}

# Fields every product row must have a non-empty value for; rows missing
# any of these are treated as blank/decorative rows and skipped.
_REQUIRED_FIELDS = ("item_code", "description")


@dataclass(frozen=True)
class ProductRow:
    ranking: str | None
    sub_group: str | None
    segment: str | None
    item_code: str
    barcode: str | None
    description: str
    color_code: str | None
    color_description: str | None
    size: str | None
    status: str | None  # raw "New / Carry over" text, e.g. "New", "Carry Over "
    retail_price: float | None
    wholesale_price: float | None
    order_qty: float | None
    order_amount: float | None
    source_sheet: str

    @property
    def sku(self) -> str:
        """A stable per-size SKU key: item code + color + size."""
        parts = [self.item_code, self.color_code or "", self.size or ""]
        return "-".join(p.strip() for p in parts if p is not None)

    def is_new_or_new_size(self) -> bool:
        """True for rows marked as a new item or a newly added size.

        The manufacturer's "New / Carry over" (or "Status", depending on the
        order form) column is free text and not consistently cased or
        spaced across seasons/brands (seen: "New", "NEW", "New sizes",
        "New colour", "New Colour", "Carry Over ", "Carry over "). We treat
        any value containing the word "new" (case-insensitive) as new;
        anything else (including blank) is treated as carry-over and
        excluded.
        """
        if not self.status:
            return False
        return "new" in self.status.strip().lower()


def _normalize_header(raw: object) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip().lower()
    text = " ".join(text.split())  # collapse repeated whitespace
    return text or None


def _find_header_row(sheet, max_scan_rows: int = 60) -> tuple[int, dict[int, str]] | None:
    """Scan the top of a sheet for the header row and return (row_index, {col_index: field_name})."""
    best_row = None
    best_map: dict[int, str] = {}
    for row_idx in range(1, min(max_scan_rows, sheet.max_row) + 1):
        col_map: dict[int, str] = {}
        for col_idx in range(1, sheet.max_column + 1):
            header = _normalize_header(sheet.cell(row=row_idx, column=col_idx).value)
            if header in _HEADER_ALIASES:
                col_map[col_idx] = _HEADER_ALIASES[header]
        # Require at least item_code + description to accept this as the header row.
        if "item_code" in col_map.values() and "description" in col_map.values():
            if len(col_map) > len(best_map):
                best_row, best_map = row_idx, col_map
    if best_row is None:
        return None
    return best_row, best_map


def _cell_str(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _cell_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_sheet(sheet, sheet_name: str) -> list[ProductRow]:
    header = _find_header_row(sheet)
    if header is None:
        return []
    header_row, col_map = header

    rows: list[ProductRow] = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        values = {
            field: sheet.cell(row=row_idx, column=col_idx).value
            for col_idx, field in col_map.items()
        }
        item_code = _cell_str(values.get("item_code"))
        description = _cell_str(values.get("description"))
        if not item_code or not description:
            continue  # blank/decorative row

        rows.append(
            ProductRow(
                ranking=_cell_str(values.get("ranking")),
                sub_group=_cell_str(values.get("sub_group")),
                segment=_cell_str(values.get("segment")),
                item_code=item_code,
                barcode=_cell_str(values.get("barcode")),
                description=description,
                color_code=_cell_str(values.get("color_code")),
                color_description=_cell_str(values.get("color_description")),
                size=_cell_str(values.get("size")),
                status=_cell_str(values.get("status")),
                retail_price=_cell_float(values.get("retail_price")),
                wholesale_price=_cell_float(values.get("wholesale_price")),
                order_qty=_cell_float(values.get("order_qty")),
                order_amount=_cell_float(values.get("order_amount")),
                source_sheet=sheet_name,
            )
        )
    return rows


def _select_default_sheets(wb) -> list[str]:
    """Pick which sheets to read when the caller doesn't name any explicitly.

    Manufacturer files we've seen bundle a single consolidated sheet (e.g.
    "North ALL products") alongside per-category sheets (kites, apparel,
    ...) that repeat the same rows. Reading everything naively double-counts
    those items. If there's exactly one sheet whose name contains "all",
    prefer it alone; otherwise fall back to every sheet with a recognizable
    product header (parse_manufacturer_file then de-duplicates by SKU as a
    safety net).
    """
    candidate_sheets = [name for name in wb.sheetnames if _find_header_row(wb[name]) is not None]
    consolidated = [name for name in candidate_sheets if "all" in name.strip().lower()]
    if len(consolidated) == 1:
        return consolidated
    return candidate_sheets


def _dedupe_by_sku(rows: list[ProductRow]) -> list[ProductRow]:
    """Drop later rows that share a SKU with an earlier one, preserving order."""
    seen: set[str] = set()
    deduped: list[ProductRow] = []
    for row in rows:
        if row.sku in seen:
            continue
        seen.add(row.sku)
        deduped.append(row)
    return deduped


def parse_manufacturer_file(
    path: str | Path,
    sheet_names: list[str] | None = None,
) -> list[ProductRow]:
    """Parse a manufacturer order-form workbook into ProductRow records.

    Args:
        path: path to the .xlsx file.
        sheet_names: optional explicit list of sheet names to read (in that
            order); returned as-is, with no de-duplication. Pass e.g.
            ["North ALL products "] to be explicit about reading only the
            consolidated sheet.
            If omitted (the normal case), sheets are auto-selected to avoid
            double-counting: a single consolidated "...all..." sheet is
            preferred if found, otherwise every sheet is read and results
            are de-duplicated by SKU (item code + color + size).
    """
    # Note: intentionally NOT read_only=True. This module does random-access
    # cell lookups (sheet.cell(row, column)), which openpyxl's read_only mode
    # handles extremely slowly (it's optimized for sequential iteration only).
    # The manufacturer files here are small (hundreds of rows), so loading
    # normally is fast and simple.
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        auto_select = sheet_names is None
        names = sheet_names if sheet_names is not None else _select_default_sheets(wb)
        all_rows: list[ProductRow] = []
        for name in names:
            if name not in wb.sheetnames:
                raise ValueError(f'Sheet "{name}" not found. Available: {wb.sheetnames}')
            all_rows.extend(parse_sheet(wb[name], name))
        return _dedupe_by_sku(all_rows) if auto_select else all_rows
    finally:
        wb.close()


def filter_new_items(rows: list[ProductRow]) -> list[ProductRow]:
    """Keep only rows that are new items or newly-added sizes."""
    return [r for r in rows if r.is_new_or_new_size()]
